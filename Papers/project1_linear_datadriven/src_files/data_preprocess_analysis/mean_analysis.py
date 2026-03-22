#!/usr/bin/env python3
"""
mean_analysis.py — Clean combined DMD + POD reconstruction error comparison
with and without mean subtraction, including alpha (fair) correction.

For each (Case, Observable) pair:
  - Preferred rank = 10.
  - If DMD mean0 doesn't have rank 10, use the lowest available rank from
    DMD mean0 and find the closest matching rank in DMD mean1.
  - POD errors are read at the same rank (POD has errors for all ranks).

Alpha correction:
  alpha_DMD = ||Sigma_m1||^2_F / ||Sigma_m0||^2_F   (from DMD SVD_S)
  alpha_POD = sum(|W_m1|) / sum(|W_m0|)              (from POD eigenvalues)
  fair_err  = raw_mean1_err * alpha

Output: mean_analysis.xlsx with one clean sheet.
"""

import glob
import os
import re
import sys
from pathlib import Path

import h5py
import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE = Path(__file__).resolve().parent.parent
DMD_DIR = BASE / "Codes" / "DMD" / "Results"
POD_DIR = BASE / "Codes" / "POD" / "Results"
OUT_DIR = Path(__file__).resolve().parent

# ---------------------------------------------------------------------------
# Filename parsing
# ---------------------------------------------------------------------------
_FNAME_RE = re.compile(
    r"Results_(?:DMD|POD)_(.+?)_(2D|3D)new_(\d+)_(\d+)_DR(\d+)_VR(\d+)"
    r"_Re_(\d+)_Bo_(\d+)_(\w+)\.h5"
)

# Variants to skip
_SKIP_VARIANTS = {"der1", "mean1test"}


def parse_filename(fpath):
    """Return dict with obs, dim, Re, Bo, variant — or None if not matched."""
    m = _FNAME_RE.search(os.path.basename(fpath))
    if m is None:
        return None
    return {
        "obs": m.group(1),
        "dim": m.group(2),
        "Re": int(m.group(7)),
        "Bo": int(m.group(8)),
        "variant": m.group(9),
    }


def canonical_obs(obs, dim):
    """Normalise POD 3D naming convention (LEGi2 → LEGi, LEUVWi2 → LEUVWi)."""
    if dim == "3D" and obs.endswith("2"):
        return obs[:-1]
    return obs


def make_key(obs, dim, Re, Bo):
    return (obs, dim, Re, Bo)


# ---------------------------------------------------------------------------
# Framework label from observable code
# ---------------------------------------------------------------------------
def framework_label(obs):
    if obs.startswith("LE"):
        return "Moving-Eulerian"
    elif obs.startswith("L"):
        return "Lagrangian"
    else:
        return "Eulerian"


# ---------------------------------------------------------------------------
# HDF5 data extraction
# ---------------------------------------------------------------------------
def extract_dmd(h5path):
    """Return (ranks, errors, total_energy) from a DMD HDF5 file.

    Corrupted Recons_vec entries (negative ranks, ranks > 10 000) are dropped.
    """
    with h5py.File(h5path, "r") as f:
        # Recons_vec: (n_evaluated, 4) → [gamma_idx, nzval/rank, gamma, error%]
        ranks = np.array([], dtype=int)
        errors = np.array([])
        if "Standard_Decomposition/Recons_vec" in f:
            rv = f["Standard_Decomposition/Recons_vec"][:]
            raw_ranks = rv[:, 1].astype(int)
            raw_errors = rv[:, 3]
            # Filter to sane values
            valid = (raw_ranks > 0) & (raw_ranks < 10_000) & np.isfinite(raw_errors)
            ranks = raw_ranks[valid]
            errors = raw_errors[valid]

        # SVD_S (r×r diagonal or 1-D)
        total_energy = np.nan
        if "Standard_Decomposition/SVD_S" in f:
            s = f["Standard_Decomposition/SVD_S"][:]
            if s.ndim == 2:
                total_energy = float(np.sum(np.diag(s) ** 2))
            else:
                total_energy = float(np.sum(s ** 2))

    return ranks, errors, total_energy


def extract_pod(h5path):
    """Return (error_vector, total_energy) from a POD HDF5 file.

    error_vector[i] = reconstruction error at rank i+1.
    total_energy = sum(|λ|), where λ are eigenvalues stored as 'Sigma'.
    """
    with h5py.File(h5path, "r") as f:
        re_err = f["Reconstruction_error"][:]
        sigma = f["Raw/Sigma"][:]
        total_energy = float(np.sum(np.abs(sigma)))
    return re_err, total_energy


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    # --- Collect files -------------------------------------------------------
    dmd_files = sorted(glob.glob(str(DMD_DIR / "Results_DMD_*.h5")))
    pod_files = sorted(glob.glob(str(POD_DIR / "Results_POD_*.h5")))

    # Nested dicts:  key → variant → data
    dmd_store = {}  # key → {mean0: ..., mean1: ...}
    pod_store = {}

    for fp in dmd_files:
        info = parse_filename(fp)
        if info is None or info["variant"] in _SKIP_VARIANTS:
            continue
        obs_c = canonical_obs(info["obs"], info["dim"])
        key = make_key(obs_c, info["dim"], info["Re"], info["Bo"])
        ranks, errors, te = extract_dmd(fp)
        dmd_store.setdefault(key, {})[info["variant"]] = {
            "ranks": ranks,
            "errors": errors,
            "total_energy": te,
            "path": fp,
        }

    for fp in pod_files:
        info = parse_filename(fp)
        if info is None or info["variant"] in _SKIP_VARIANTS:
            continue
        obs_c = canonical_obs(info["obs"], info["dim"])
        key = make_key(obs_c, info["dim"], info["Re"], info["Bo"])
        errs, te = extract_pod(fp)
        pod_store.setdefault(key, {})[info["variant"]] = {
            "errors": errs,
            "total_energy": te,
            "path": fp,
        }

    # --- Build rows ----------------------------------------------------------
    TARGET_RANK = 10
    rows = []
    all_keys = sorted(set(dmd_store.keys()) | set(pod_store.keys()))

    for key in all_keys:
        obs, dim, Re, Bo = key
        d_m0 = dmd_store.get(key, {}).get("mean0")
        d_m1 = dmd_store.get(key, {}).get("mean1")
        p_m0 = pod_store.get(key, {}).get("mean0")
        p_m1 = pod_store.get(key, {}).get("mean1")

        # ---- Rank selection (driven by DMD mean0) ----------------------------
        rank = TARGET_RANK
        dmd_m0_err = np.nan
        dmd_m1_err = np.nan
        dmd_m1_rank = np.nan

        has_m0_ranks = d_m0 is not None and len(d_m0["ranks"]) > 0
        has_m1_ranks = d_m1 is not None and len(d_m1["ranks"]) > 0

        if has_m0_ranks:
            if TARGET_RANK in d_m0["ranks"]:
                idx = int(np.where(d_m0["ranks"] == TARGET_RANK)[0][0])
                dmd_m0_err = d_m0["errors"][idx]
                rank = TARGET_RANK
            else:
                # Closest available rank to TARGET in DMD mean0
                closest_idx = int(np.argmin(np.abs(d_m0["ranks"] - TARGET_RANK)))
                rank = int(d_m0["ranks"][closest_idx])
                dmd_m0_err = d_m0["errors"][closest_idx]
        elif has_m1_ranks:
            # No valid mean0 Recons_vec — fall back to mean1's rank 10
            if TARGET_RANK in d_m1["ranks"]:
                rank = TARGET_RANK
            else:
                closest_idx = int(np.argmin(np.abs(d_m1["ranks"] - TARGET_RANK)))
                rank = int(d_m1["ranks"][closest_idx])

        if has_m1_ranks:
            # Find rank closest to `rank`
            closest_idx = int(np.argmin(np.abs(d_m1["ranks"] - rank)))
            dmd_m1_rank = int(d_m1["ranks"][closest_idx])
            dmd_m1_err = d_m1["errors"][closest_idx]

        # ---- POD errors at same rank -----------------------------------------
        pod_m0_err = np.nan
        pod_m1_err = np.nan
        if p_m0 is not None and rank <= len(p_m0["errors"]):
            pod_m0_err = p_m0["errors"][rank - 1]
        if p_m1 is not None and rank <= len(p_m1["errors"]):
            pod_m1_err = p_m1["errors"][rank - 1]

        # ---- Alpha & fair errors ---------------------------------------------
        alpha_dmd = np.nan
        if d_m0 is not None and d_m1 is not None:
            alpha_dmd = d_m1["total_energy"] / d_m0["total_energy"]

        alpha_pod = np.nan
        if p_m0 is not None and p_m1 is not None:
            alpha_pod = p_m1["total_energy"] / p_m0["total_energy"]

        dmd_m1_fair = (
            dmd_m1_err * alpha_dmd
            if not (np.isnan(dmd_m1_err) or np.isnan(alpha_dmd))
            else np.nan
        )
        pod_m1_fair = (
            pod_m1_err * alpha_pod
            if not (np.isnan(pod_m1_err) or np.isnan(alpha_pod))
            else np.nan
        )

        rows.append(
            {
                "Case": f"Re{Re}Bo{Bo}",
                "Framework": framework_label(obs),
                "Observable": obs,
                "Dim": dim,
                "Rank": rank,
                "DMD_m1_rank": int(dmd_m1_rank) if not np.isnan(dmd_m1_rank) else "",
                "DMD_m0_err%": round_or_nan(dmd_m0_err),
                "DMD_m1_err%": round_or_nan(dmd_m1_err),
                "DMD_m1_fair%": round_or_nan(dmd_m1_fair),
                "POD_m0_err%": round_or_nan(pod_m0_err),
                "POD_m1_err%": round_or_nan(pod_m1_err),
                "POD_m1_fair%": round_or_nan(pod_m1_fair),
                "α_DMD": round_or_nan(alpha_dmd, 6),
                "α_POD": round_or_nan(alpha_pod, 6),
            }
        )

    df = pd.DataFrame(rows)

    # --- Sort: Eulerian → ME → Lagrangian, then Case --------------------------
    fw_order = {"Eulerian": 0, "Moving-Eulerian": 1, "Lagrangian": 2}
    df["_fw_sort"] = df["Framework"].map(fw_order)
    df = df.sort_values(["_fw_sort", "Case", "Observable"]).drop(columns="_fw_sort")
    df.reset_index(drop=True, inplace=True)

    # --- Print to console -----------------------------------------------------
    print(df.to_string(index=False))
    print(f"\nTotal rows: {len(df)}")

    # --- Export ---------------------------------------------------------------
    xlsx_path = OUT_DIR / "mean_analysis.xlsx"
    csv_path = OUT_DIR / "mean_analysis.csv"
    df.to_csv(csv_path, index=False)
    print(f"\nCSV  → {csv_path}")

    if HAS_OPENPYXL:
        export_excel(df, xlsx_path)
        print(f"XLSX → {xlsx_path}")
    else:
        print("openpyxl not available — skipping Excel export")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def round_or_nan(val, decimals=4):
    if isinstance(val, float) and np.isnan(val):
        return ""
    return round(float(val), decimals)


def export_excel(df, xlsx_path):
    """Write a clean, formatted Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mean Analysis"

    # ---- Styles ----
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    num_fmt_4 = "0.0000"
    num_fmt_6 = "0.000000"

    # ---- Write header ----
    headers = list(df.columns)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ---- Write data ----
    for r, row_data in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val if val != "" else None)
            cell.alignment = center_align
            cell.border = thin_border
            col_name = headers[c - 1]
            if "err%" in col_name or "fair%" in col_name:
                cell.number_format = num_fmt_4
            elif col_name.startswith("α"):
                cell.number_format = num_fmt_6

    # ---- Column widths ----
    col_widths = {
        "Case": 12,
        "Framework": 18,
        "Observable": 12,
        "Dim": 5,
        "Rank": 6,
        "DMD_m1_rank": 12,
        "DMD_m0_err%": 14,
        "DMD_m1_err%": 14,
        "DMD_m1_fair%": 14,
        "POD_m0_err%": 14,
        "POD_m1_err%": 14,
        "POD_m1_fair%": 14,
        "α_DMD": 10,
        "α_POD": 10,
    }
    for c, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c)].width = col_widths.get(h, 12)

    # ---- Freeze top row ----
    ws.freeze_panes = "A2"

    # ---- Auto-filter ----
    ws.auto_filter.ref = ws.dimensions

    wb.save(xlsx_path)


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    main()
